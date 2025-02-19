import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.chart.axis import DateAxis

# Create the data dictionary
data = {
    "Year": [2010, 2020, 2030, 2040, 2050, 2060, 2070, 2080, 2090, 2100],
    "Sub-Saharan Africa": [200, 400, 600, 900, 1200, 1600, 2000, 2400, 2800, 3100],
    "East Asia & Pacific": [800, 1600, 1900, 2000, 2050, 2000, 1950, 1900, 1900, 1900],
    "Europe & Central Asia": [400, 500, 700, 900, 1100, 1300, 1500, 1700, 1850, 1950],
    "Latin America & Caribbean": [500, 900, 1000, 1100, 1200, 1250, 1300, 1300, 1350, 1350],
    "Middle East and North Africa": [200, 300, 400, 500, 600, 650, 700, 750, 800, 850],
    "South Asia": [300, 500, 700, 850, 1000, 1200, 1400, 1600, 1800, 1900],
    "High Income & OECD": [1700, 1850, 1900, 1950, 2000, 2000, 1950, 1900, 1900, 1900]
}

# Convert to DataFrame
df = pd.DataFrame(data)

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Waste Generation Data"

# Add title
ws['A1'] = 'Regional Waste Generation Projections (thousand tonnes/day)'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:H1')
ws['A1'].alignment = Alignment(horizontal='center')

# Write the DataFrame to Excel, starting from row 2 to accommodate the title
df.to_excel(wb, sheet_name='Waste Generation Data', startrow=1, index=False)

# Create the line chart
chart = LineChart()
chart.title = "Waste Generation Projections by Region"
chart.style = 2
chart.height = 15  # Height in cm
chart.width = 25   # Width in cm

# Configure the data
data = Reference(ws, min_row=2, max_row=12, min_col=1, max_col=8)
cats = Reference(ws, min_row=3, max_row=12, min_col=1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Customize chart appearance
chart.x_axis.title = "Year"
chart.y_axis.title = "Waste Generation (thousand tonnes/day)"
chart.legend.position = 'b'  # Position legend at bottom

# Add the chart to the worksheet
ws.add_chart(chart, "J3")

# Save the workbook
wb.save('waste_generation_projections.xlsx')